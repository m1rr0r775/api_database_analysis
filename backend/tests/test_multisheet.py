import os
import tempfile
import unittest

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.main import app


def _make_xlsx(path: str) -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "订单主表"
    ws1.append(["订单ID", "客户ID", "金额"])
    ws1.append([1, "c1", 10])
    ws1.append([2, "c2", 20])

    ws2 = wb.create_sheet("订单明细表")
    ws2.append(["订单ID", "商品", "数量"])
    ws2.append([1, "A", 2])
    ws2.append([1, "B", 1])
    ws2.append([2, "A", 3])

    ws3 = wb.create_sheet("模板")
    ws3.append(["说明"])
    ws3.append(["这是模板Sheet"])

    ws4 = wb.create_sheet("1月销售")
    ws4.append(["日期", "销售额"])
    ws4.append(["2025-01-01", 100])
    ws4.append(["2025-01-02", 120])

    ws5 = wb.create_sheet("2月销售")
    ws5.append(["日期", "销售额"])
    ws5.append(["2025-02-01", 90])
    ws5.append(["2025-02-02", 110])

    wb.save(path)


class TestMultiSheet(unittest.TestCase):
    def test_upload_list_preview_extract_and_join(self):
        c = TestClient(app)
        sid = c.post("/api/sessions/").json()["session_id"]

        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        try:
            _make_xlsx(tmp.name)
            with open(tmp.name, "rb") as f:
                resp = c.post(
                    f"/api/sessions/{sid}/files/",
                    files=[("files", ("multi.xlsx", f.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))],
                )
            self.assertEqual(resp.status_code, 200, resp.text)
            files = resp.json()["files"]
            self.assertEqual(len(files), 1)
            wb_entry = files[0]
            self.assertEqual(wb_entry.get("kind"), "excel_workbook")
            wb_id = wb_entry["file_id"]

            ls = c.get(f"/api/sessions/{sid}/excel/{wb_id}/sheets/")
            self.assertEqual(ls.status_code, 200, ls.text)
            sheets = ls.json()["sheets"]
            self.assertTrue(any(s["name"] == "订单主表" for s in sheets))
            self.assertTrue(any(s["name"] == "订单明细表" for s in sheets))
            self.assertTrue(any(s["name"] == "模板" for s in sheets))

            pv = c.post(f"/api/sessions/{sid}/excel/{wb_id}/preview/", json={"sheet": "订单主表", "smart_clean": True, "options": {}})
            self.assertEqual(pv.status_code, 200, pv.text)
            self.assertGreaterEqual((pv.json().get("cleaned") or {}).get("row_count", 0), 2)

            ex = c.post(
                f"/api/sessions/{sid}/excel/{wb_id}/extract/",
                json={"sheets": ["订单主表", "订单明细表"], "smart_clean": True, "options": {}, "stack_similar": False},
            )
            self.assertEqual(ex.status_code, 200, ex.text)
            extracted = ex.json()["files"]
            self.assertEqual(len(extracted), 2)
            left_id = extracted[0]["file_id"]
            right_id = extracted[1]["file_id"]

            sug = c.post(f"/api/sessions/{sid}/model/suggest/", json={"file_ids": [left_id, right_id]})
            self.assertEqual(sug.status_code, 200, sug.text)
            self.assertTrue(len(sug.json().get("suggestions") or []) >= 1)

            b = c.post(
                f"/api/sessions/{sid}/model/build/",
                json={
                    "left_file_id": left_id,
                    "right_file_id": right_id,
                    "left_key": "订单ID",
                    "right_key": "订单ID",
                    "how": "left",
                    "name": "joined",
                },
            )
            self.assertEqual(b.status_code, 200, b.text)
            merged = b.json()["file"]
            self.assertTrue(merged.get("row_count", 0) > 0)
            self.assertTrue("订单ID" in (merged.get("columns") or []))

            ex2 = c.post(
                f"/api/sessions/{sid}/excel/{wb_id}/extract/",
                json={"sheets": ["1月销售", "2月销售"], "smart_clean": True, "options": {}, "stack_similar": True},
            )
            self.assertEqual(ex2.status_code, 200, ex2.text)
            stacked = ex2.json()["files"]
            self.assertTrue(len(stacked) >= 1)
            any_stacked = any(f.get("source") == "excel_stacked" for f in stacked)
            self.assertTrue(any_stacked)
            stacked_entry = next((f for f in stacked if f.get("source") == "excel_stacked"), None)
            self.assertIsNotNone(stacked_entry)
            self.assertTrue("__sheet__" in (stacked_entry.get("columns") or []))
        finally:
            try:
                os.remove(tmp.name)
            except Exception:
                pass
