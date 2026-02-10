import os
import tempfile
import unittest

from openpyxl import load_workbook

from app.services.table_io import read_table
from app.services.excel_exporter import workbook_for_chart


class TestTableIo(unittest.TestCase):
    def test_read_table_dedupes_columns(self):
        fd, path = tempfile.mkstemp(suffix=".csv")
        os.close(fd)
        try:
            with open(path, "w", encoding="utf-8") as f:
                f.write("a,a\n1,2\n")
            df = read_table(path)
            self.assertEqual(len(df.columns), 2)
            self.assertNotEqual(df.columns[0], df.columns[1])
        finally:
            try:
                os.remove(path)
            except Exception:
                pass


class TestExcelExport(unittest.TestCase):
    def test_excel_formula_injection_is_escaped(self):
        option = {
            "xAxis": {"type": "category", "data": ["=1+1", "ok"]},
            "yAxis": {"type": "value"},
            "series": [{"type": "bar", "name": "s", "data": [1, 2]}],
        }
        blob, _ = workbook_for_chart("t", option)
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        try:
            with open(tmp.name, "wb") as f:
                f.write(blob)
            wb = load_workbook(tmp.name, data_only=False)
            ws = wb.active
            self.assertTrue(str(ws["A2"].value).startswith("'="))
        finally:
            try:
                os.remove(tmp.name)
            except Exception:
                pass
