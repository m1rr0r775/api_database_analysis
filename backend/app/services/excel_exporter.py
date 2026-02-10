from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import AreaChart, BarChart, LineChart, PieChart, ScatterChart, Reference, Series
from openpyxl.chart.label import DataLabelList


def _safe_excel_text(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    s = value.replace("\x00", "")
    if not s:
        return s
    if s[0] in ("=", "+", "-", "@"):
        return "'" + s
    return s


def _safe_sheet_title(title: str, fallback: str) -> str:
    t = (title or "").strip()
    if not t:
        t = fallback
    t = re.sub(r"[\[\]\*\?:/\\]", "_", t)
    return t[:31]


def _safe_filename(title: str, fallback: str) -> str:
    name = (title or "").strip() or fallback
    name = re.sub(r"[^\w\u4e00-\u9fff\- ]+", "_", name).strip()
    name = re.sub(r"\s+", "_", name)
    return (name[:80] or fallback) + ".xlsx"


def _is_series(option: dict, kind: str) -> bool:
    s = (option.get("series") or [])
    if not s or not isinstance(s, list):
        return False
    first = s[0] if isinstance(s[0], dict) else None
    return bool(first and str(first.get("type", "")).lower() == kind)


def workbook_for_chart(title: str, option: dict[str, Any]) -> tuple[bytes, str]:
    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_title(title, "Chart")

    _write_chart_sheet(ws, title=title, option=option)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue(), _safe_filename(title, "chart_export")


def workbook_for_dashboard(title: str, charts: list[dict[str, Any]]) -> tuple[bytes, str]:
    wb = Workbook()
    wb.remove(wb.active)

    for idx, c in enumerate(charts, start=1):
        chart_title = str(c.get("title") or f"Chart {idx}")
        option = c.get("option") or {}
        ws = wb.create_sheet(_safe_sheet_title(chart_title, f"Chart{idx}"))
        _write_chart_sheet(ws, title=chart_title, option=option)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue(), _safe_filename(title, "dashboard_export")


def _write_chart_sheet(ws, title: str, option: dict[str, Any]) -> None:
    if _is_series(option, "pie"):
        _write_pie(ws, title, option)
        return
    if _is_series(option, "scatter"):
        _write_scatter(ws, title, option)
        return
    if _is_series(option, "boxplot"):
        _write_boxplot_data_only(ws, title, option)
        return

    chart_kind = _extract_cartesian_kind(option)
    if chart_kind == "line":
        _write_cartesian(ws, title, option, kind="line")
        return
    if chart_kind == "area":
        _write_cartesian(ws, title, option, kind="area")
        return

    _write_cartesian(ws, title, option, kind="bar")


def _extract_cartesian_kind(option: dict[str, Any]) -> str:
    series = option.get("series") or []
    if not isinstance(series, list) or not series:
        return "bar"
    first = series[0] if isinstance(series[0], dict) else {}
    t = str(first.get("type", "")).lower()
    if t in ("line", "bar"):
        return t
    if t in ("scatter", "pie", "boxplot"):
        return t
    if t == "bar" and first.get("stack"):
        return "bar"
    if first.get("areaStyle") is not None:
        return "area"
    return "bar"


def _write_cartesian(ws, title: str, option: dict[str, Any], kind: str) -> None:
    x_axis = option.get("xAxis") or {}
    if isinstance(x_axis, list):
        x_axis = x_axis[0] if x_axis else {}
    categories = (x_axis.get("data") or [])
    categories = [str(c) for c in categories]

    series = option.get("series") or []
    if not isinstance(series, list):
        series = []

    ws["A1"] = "Category"
    for i, cat in enumerate(categories, start=2):
        ws.cell(row=i, column=1, value=_safe_excel_text(cat))

    for s_idx, s in enumerate(series, start=1):
        if not isinstance(s, dict):
            continue
        name = str(s.get("name") or f"Series{s_idx}")
        ws.cell(row=1, column=1 + s_idx, value=_safe_excel_text(name))
        data = s.get("data") or []
        for i, v in enumerate(data, start=2):
            ws.cell(row=i, column=1 + s_idx, value=v)

    max_row = 1 + len(categories)
    max_col = 1 + max(1, len(series))

    if kind == "line":
        chart = LineChart()
    elif kind == "area":
        chart = AreaChart()
    else:
        chart = BarChart()
        first = series[0] if series and isinstance(series[0], dict) else {}
        if first.get("stack"):
            chart.grouping = "stacked"

    chart.title = title or "Chart"
    chart.legend = None if len(series) <= 1 else chart.legend
    chart.style = 10

    data_ref = Reference(ws, min_col=2, min_row=1, max_col=max_col, max_row=max_row)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=max_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = False

    ws.add_chart(chart, "E2")


def _write_pie(ws, title: str, option: dict[str, Any]) -> None:
    series = option.get("series") or []
    first = series[0] if series and isinstance(series[0], dict) else {}
    data = first.get("data") or []
    if not isinstance(data, list):
        data = []

    ws["A1"] = "Category"
    ws["B1"] = "Value"
    for i, item in enumerate(data, start=2):
        if not isinstance(item, dict):
            continue
        ws.cell(row=i, column=1, value=_safe_excel_text(str(item.get("name", ""))))
        ws.cell(row=i, column=2, value=item.get("value"))

    last_row = 1 + len(data)
    chart = PieChart()
    chart.title = title or "Pie"
    values = Reference(ws, min_col=2, min_row=1, max_row=last_row)
    labels = Reference(ws, min_col=1, min_row=2, max_row=last_row)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(labels)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showVal = False
    ws.add_chart(chart, "D2")


def _write_scatter(ws, title: str, option: dict[str, Any]) -> None:
    series = option.get("series") or []
    if not isinstance(series, list) or not series:
        return

    chart = ScatterChart()
    chart.title = title or "Scatter"
    chart.style = 2

    col = 1
    for s_idx, s in enumerate(series, start=1):
        if not isinstance(s, dict):
            continue
        name = str(s.get("name") or f"Series{s_idx}")
        points = s.get("data") or []
        if not isinstance(points, list) or not points:
            continue

        ws.cell(row=1, column=col, value=_safe_excel_text(f"{name}_x"))
        ws.cell(row=1, column=col + 1, value=_safe_excel_text(f"{name}_y"))
        row = 2
        for p in points:
            if not isinstance(p, list) or len(p) < 2:
                continue
            ws.cell(row=row, column=col, value=p[0])
            ws.cell(row=row, column=col + 1, value=p[1])
            row += 1

        xvalues = Reference(ws, min_col=col, min_row=2, max_row=row - 1)
        yvalues = Reference(ws, min_col=col + 1, min_row=2, max_row=row - 1)
        chart.series.append(Series(yvalues, xvalues, title=name))
        col += 3

    ws.add_chart(chart, "H2")


def _write_boxplot_data_only(ws, title: str, option: dict[str, Any]) -> None:
    ws["A1"] = "Series"
    ws["B1"] = "Min"
    ws["C1"] = "Q1"
    ws["D1"] = "Median"
    ws["E1"] = "Q3"
    ws["F1"] = "Max"

    series = option.get("series") or []
    first = series[0] if series and isinstance(series[0], dict) else {}
    data = first.get("data") or []
    if not isinstance(data, list):
        data = []

    for i, row in enumerate(data, start=2):
        ws.cell(row=i, column=1, value=_safe_excel_text(f"{title}_{i-1}"))
        if isinstance(row, list) and len(row) >= 5:
            for j in range(5):
                ws.cell(row=i, column=2 + j, value=row[j])
